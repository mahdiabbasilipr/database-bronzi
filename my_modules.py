import openpyxl
from product_class import Product

# i wan't to write a database on excel file

# function goto get input from user and return that in string that can change to dict by eval()
def get_specifications():
    # get Specifications of new product and append them look like a dictonary
    print("we need some information about your product")

    while (True):  # FOR ERROR CONTROL code VALUE ERROR
        try:
            code = int(input('plese enter code: '))

        except ValueError:
            print("please enter a integer Value")

        else:
            break

    name = input('please enter name: ')  # name is string and don't need error control value  error

    while (True):  # FOR ERROR CONTROL weight VALUE ERROR
        try:
            weight = input('please enter weight(kilo): ')

        except ValueError:
            print("please enter a integer Value")

        else:
            break

    while (True):  # FOR ERROR CONTROL hight VALUE ERROR
        try:
            hight = input('please enter hight(centimeter): ')

        except ValueError:
            print("please enter a integer Value")

        else:
            break

    while (True):  # FOR ERROR CONTROL width VALUE ERROR
        try:
            width = input('please enter width(centimeter): ')

        except ValueError:
            print("please enter a integer Value")

        else:
            break

    while (True):  # FOR ERROR CONTROL price VALUE ERROR
        try:
            price = int(input('please enter price(integer and toman): '))

        except ValueError:
            print("please enter a integer Value")

        else:
            break

    while (True):  # FOR ERROR CONTROL price VALUE ERROR
        try:
            price_of_bronze = int(input('please enter price of bronze of today:(if you dont know enter 0) '))

        except ValueError:
            print("please enter a integer Value")

        else:
            break
    # print data for get user approval
    print(
        "\nyou\'r new product is :\nname = %s \nweight = %s \nhight = %s \nwidth = %s \nprice = %s" % (
            name, weight, hight, width, price
        )
    )
    user_approval = input("\nare all of the information that enter ,True??:(y/n) ")
    if user_approval == 'n':  # User approval
        final_dict = get_specifications()  # run again
    else:
        # make product object
        new_product = Product(name, code, weight, hight, width, price)

    return new_product.get_dict_data()

# get list data is a function to make a list of dictionary data
def get_list_data(addres_of_excel):
    wb = openpyxl.load_workbook(filename=addres_of_excel)  # make work book
    sheet_range = wb.get_sheet_by_name('sheet1')  # get sheet1 to calc max row
    max_row = sheet_range.max_row
    list_data = []
    for i in range(2, max_row + 1):
        i = str(i)
        new_dict = {
            'name': sheet_range['A' + i].value,
            'code': sheet_range['B' + i].value,
            'weight': sheet_range['C' + i].value,
            'hight': sheet_range['D' + i].value,
            'width': sheet_range['E' + i].value,
            'price': sheet_range['F' + i].value,
            'price_for_customer': sheet_range['H' + i].value
        }
        list_data.append(new_dict)
        i = int(i)
    return list_data


# write excel is function that i use for write or over write excel
def write_excel(addres_of_file, list_data):
    wb = openpyxl.Workbook()
    ws1 = wb.create_sheet('sheet1')

    ws1['A1'] = 'نام محصول'
    ws1['B1'] = 'کد محصول'
    ws1['C1'] = 'وزن'
    ws1['D1'] = 'قد'
    ws1['E1'] = 'عرض'
    ws1['F1'] = 'قیمت برای ما'
    ws1['G1'] = 'قیمت برای مشتری'

    temp = 1
    for item in list_data:
        temp += 1
        temp = str(temp)

        ws1['A' + temp] = item['name']
        ws1['B' + temp] = item['code']
        ws1['C' + temp] = item['weight']
        ws1['D' + temp] = item['hight']
        ws1['E' + temp] = item['width']
        ws1['F' + temp] = item['price']
        ws1['G' + temp] = item['price_for_customer']

        temp = int(temp)

    wb.save(filename=addres_of_file)


def append_new_data(addres_of_file):
    new_data = get_specifications()
    list_data = get_list_data(addres_of_file)
    list_data.append(new_data)
    write_excel(addres_of_file, list_data)
